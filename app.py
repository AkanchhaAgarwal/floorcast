"""
Floorcast — seat planning for a contact centre estate.

Built to mirror how a seat planner actually works: a portfolio of floors across
several countries, seats in four states (allocated, available, trapped,
expansion), demand assembled per site and per period, ramps matched floor by
floor, and only then a coloured floor map.

Run:  streamlit run app.py
"""

import io
import os as _os
import numpy as np
import pandas as pd
import streamlit as st

from floorcast_core import estate as es
from floorcast_core import demand_reader as dr
from floorcast_core import seatmap_engine as sm
from floorcast_core import floor_render as fr
from floorcast_core import moves as mv
from floorcast_core import restrictions as rx
from floorcast_core import scenario as sc
from floorcast_core import thresholds as th
from floorcast_core import commercial as cm
from floorcast_core import expansion as ex
from floorcast_core import oneview as ov
from floorcast_core import sharing as sg
from floorcast_core import seating as se
from floorcast_core import revenue as rv
from floorcast_core import onesource as os1
from floorcast_core import roles as rl
from floorcast_core import plan_library as pl

st.set_page_config(page_title="Floorcast", page_icon="🏬", layout="wide")

TEMPLATES = [
    ("Floor inventory", "1 · required", "floor_inventory.csv", "csv",
     ["geo", "country", "city", "site", "building", "floor", "total_seats", "allocated",
      "available", "trapped", "trapped_reason", "expansion_space", "expansion_eta_weeks",
      "programs", "notes"],
     "One row per floor — seats split into allocated, available and unusable"),
    ("Planning workbook", "2 · required", "planning_workbook.xlsx", "xlsx",
     ["Account", "LOB", "Country", "City", "Site", "Metric"],
     "Account x LOB x site, a metric row each, one column per period"),
    ("Deals not yet won", "3 · optional", "deals_not_yet_won.csv", "csv",
     ["account", "opportunity", "country", "city", "site", "stage", "probability",
      "month", "hc"],
     "New business being chased, with how likely each is to close"),
    ("Allocations by floor", "4 · optional", "allocations.csv", "csv",
     ["site", "building", "floor", "account", "lob", "seats"],
     "Who holds which seats — unlocks consolidation and relocation"),
    ("Restrictions", "5 · optional", "restrictions.csv", "csv",
     ["rule", "subject", "object", "note"],
     "frozen · dedicated · no_colocate · requires · max_moves"),
]


def _template_bytes(kind, cols):
    """A blank file with the right headers. A new account needs a template, not
    somebody else's numbers to delete."""
    if kind == "csv":
        return (",".join(cols) + "\n").encode()
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook(); ws = wb.active; ws.title = "Demand"
    ws.append([None] * len(cols) + ["2026-09-01", "2026-10-01", "2026-11-01"])
    ws.append(cols + [None, None, None])
    for c in ws[2]:
        if c.value:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="1F3864")
    for metric in ("HC Forecast", "Seat Ratio", "Seats Required"):
        ws.append([None] * (len(cols) - 1) + [metric])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


@st.cache_data(show_spinner=False)
def sample_bundle() -> bytes:
    import zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for name, tag, fname, kind, cols, desc in TEMPLATES:
            z.writestr(fname, _template_bytes(kind, cols))
        z.writestr("README.txt",
                   "Floorcast input templates\n=========================\n\n"
                   "Fill these in and upload them from the sidebar. Keep the column "
                   "names; the order does not matter.\n\n"
                   + "\n".join(f"{n:22} {f:26} {d}" for n, t, f, k, c, d in TEMPLATES))
    return buf.getvalue()


def _status_style(df, col="status"):
    """Colour the status column so a region's state reads at a glance."""
    if df is None or df.empty or col not in df.columns:
        return df
    shade = {"Critical": "background-color:#FBEDEC;color:#B3261E;font-weight:600",
             "Strained": "background-color:#FBF6EC;color:#8A6D1F;font-weight:600",
             "Healthy": "background-color:#EAF4EE;color:#1E7A46"}
    return df.style.map(lambda v: shade.get(v, ""), subset=[col])


def _xlsx(sheets: dict) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as xw:
        for nm, df_ in sheets.items():
            if df_ is not None and not df_.empty:
                df_.to_excel(xw, sheet_name=nm[:31], index=False)
    return buf.getvalue()


st.markdown(
    "<h1 style='color:#0F6B6B;margin-bottom:0'>Floorcast</h1>"
    "<p style='color:#C9962E;font-weight:600;margin-top:0'>"
    "Seat planning across the estate — capacity, ramps and floor maps in one place.</p>",
    unsafe_allow_html=True)


# ─────────────────────────────────────────────── loaders
@st.cache_data(show_spinner=False)
def load_floors(data):
    return pd.read_csv(io.BytesIO(data))


@st.cache_data(show_spinner=False)
def load_demand(data):
    return dr.read_demand(io.BytesIO(data))


@st.cache_data(show_spinner=False)
def load_pipeline(data):
    return dr.read_pipeline(io.BytesIO(data))


@st.cache_data(show_spinner=False)
def load_alloc(data):
    return pd.read_csv(io.BytesIO(data))


@st.cache_data(show_spinner=False)
def load_rules(data):
    return pd.read_csv(io.BytesIO(data))


@st.cache_data(show_spinner="Reading the floor plans…")
def build_library(payloads):
    """payloads: tuple of (filename, bytes). Returns the plan dict plus anything
    that could not be identified, so the app can ask rather than guess."""
    from floorcast_core import plan_ingest
    plans, unknown = {}, []
    for name, data in payloads:
        try:
            if name.lower().endswith(".csv"):
                seats = pd.read_csv(io.BytesIO(data))
                out = {"seats": seats, "background": None, "extent": None,
                       "scale_mm_per_pt": None, "zone_labels": pd.DataFrame()}
            else:
                out = plan_ingest.read_plan(data)
        except Exception as exc:
            unknown.append({"file": name, "why": str(exc)[:120]})
            continue
        k = pl.identify(name, out)
        if not k["site"] or not k["floor"]:
            unknown.append({"file": name, "why": "could not tell which floor this is"})
            continue
        building = k["site"].split("-")[1] + " Tower" if "-" in k["site"] else "B1"
        plans[(k["site"], building, k["floor"])] = out
    return plans, unknown


@st.cache_data(show_spinner="Reading the floor plan…")
def load_plan(data, name):
    if name.lower().endswith(".csv"):
        return {"seats": pd.read_csv(io.BytesIO(data)), "background": None,
                "extent": None, "scale_mm_per_pt": None, "source": "csv"}
    from floorcast_core import plan_ingest
    out = plan_ingest.read_plan(data)
    out["source"] = "pdf"
    return out


with st.sidebar:
    st.header("Inputs")
    st.caption("Upload your files. Nothing is kept between sessions.")
    seat_model = st.radio("How are seats allocated?", list(se.MODELS),
                          format_func=lambda m: se.MODELS[m]["label"], key="seat_model",
                          help="Pooled for contact centres, assigned where everyone owns a "
                               "desk, neighbourhoods where a team owns a zone.")
    st.caption(se.MODELS[seat_model]["for"] + ". " + se.MODELS[seat_model]["ratio"] + ".")
    st.divider()
    if True:
        f_file = st.file_uploader("1 · Floor inventory (CSV)", type="csv",
                                  help="geo, country, city, site, building, floor, "
                                       "total_seats, allocated, available, trapped, "
                                       "trapped_reason, expansion_space, expansion_eta_weeks")
        d_file = st.file_uploader("2 · Planning workbook (Excel)", type=["xlsx", "xls"],
                                  help="Account / LOB / Site with metric rows and one "
                                       "column per period")
        p_file = st.file_uploader("3 · Deals not yet won (CSV) — optional", type="csv",
                                  help="account, site, stage, probability, month, hc — one row per deal")
        a_file = st.file_uploader("4 · Allocations by floor (CSV) — optional", type="csv",
                                  help="site, building, floor, account, lob, seats — who holds "
                                       "which seats. Unlocks consolidation and relocation options.")
        sh_file = st.file_uploader("8 · Shift patterns (CSV) — optional", type="csv",
                                   help="account, site, shift, start_hour, end_hour, "
                                        "system_locked — lets two accounts on opposite "
                                        "shifts share a seat")
        rt_file = st.file_uploader("9 · Rate card (CSV) — optional", type="csv",
                                   help="site, rate_per_seat (and optionally account) — "
                                        "turns the plan into revenue per seat")
        ro_file = st.file_uploader("10 · Employee roster (CSV) — optional", type="csv",
                                   help="employee_id, account, site (and seat_id if they "
                                        "already have one) — for assigned seating")
        s_file = st.file_uploader("6 · Programmes winding down (CSV) — optional", type="csv",
                                  help="account, lob, site, end_month, reason, "
                                       "seats_released — demand that is leaving")
        r_file = st.file_uploader("5 · Restrictions (CSV) — optional", type="csv",
                                  help="rule, subject, object, note — frozen accounts, "
                                       "dedicated floors, no-colocation pairs, move ceiling")
        plan_file = st.file_uploader("7 · Floor plans (PDF) — optional",
                                     type=["pdf", "csv"], accept_multiple_files=True,
                                     help="One vector PDF per surveyed floor. Upload the whole "
                                          "site and demand is placed across its floors at once. "
                                          "Name files like SITE-CODE_05F.pdf, or the drawing's "
                                          "own title block is read.")
        st.caption("The two required files get you a working plan.")

    st.divider()
    with st.expander("⬇ Blank templates", expanded=True):
        st.caption("Empty files with the right column names.")
        st.download_button("📦 All templates (.zip)", sample_bundle(),
                           file_name="floorcast_templates.zip", mime="application/zip",
                           width="stretch", key="dl_bundle_side")
        for name, tag, fname, kind, cols, desc in TEMPLATES:
            st.download_button(name, _template_bytes(kind, cols), file_name=fname,
                               mime="text/csv" if kind == "csv" else
                               "application/vnd.openxmlformats-officedocument."
                               "spreadsheetml.sheet",
                               width="stretch", key="dl_" + fname)

def _bundled(path, loader):
    """Load a bundled sample if one happens to be present. Nothing is bundled by
    default — the app starts empty, the way a new account does."""
    try:
        with open(path, "rb") as fh:
            return loader(fh.read())
    except FileNotFoundError:
        return None


floors_raw = load_floors(f_file.getvalue()) if f_file \
    else _bundled("data/sample_estate_floors.csv", load_floors)
demand = load_demand(d_file.getvalue()) if d_file \
    else _bundled("data/sample_estate_demand.xlsx", load_demand)
pipe = load_pipeline(p_file.getvalue()) if p_file \
    else _bundled("data/sample_pipeline.csv", load_pipeline)
alloc_raw = load_alloc(a_file.getvalue()) if a_file \
    else _bundled("data/sample_allocations.csv", load_alloc)
rules_raw = load_rules(r_file.getvalue()) if r_file \
    else _bundled("data/sample_restrictions.csv", load_rules)
shifts_raw = load_alloc(sh_file.getvalue()) if sh_file \
    else _bundled("data/sample_shifts.csv", load_alloc)
rates_raw = load_alloc(rt_file.getvalue()) if rt_file \
    else _bundled("data/sample_rate_card.csv", load_alloc)
roster_raw = load_alloc(ro_file.getvalue()) if ro_file \
    else _bundled("data/sample_roster.csv", load_alloc)
sunset_raw = load_alloc(s_file.getvalue()) if s_file \
    else _bundled("data/sample_sunsetting.csv", load_alloc)
if sunset_raw is None:
    sunset_raw = pd.DataFrame(columns=["account", "lob", "site", "end_month",
                                       "reason", "seats_released"])
if pipe is None:
    pipe = pd.DataFrame(columns=["account", "site", "probability", "month", "hc"])
if alloc_raw is None:
    alloc_raw = pd.DataFrame(columns=["site", "building", "floor", "account", "lob", "seats"])
if rules_raw is None:
    rules_raw = rx.empty()

# ── nothing loaded yet: show the setup screen and stop
if floors_raw is None or demand is None:
    from floorcast_core import onboarding as ob
    have = {"floors": 0 if floors_raw is None else len(floors_raw),
            "demand": 0 if demand is None else len(demand),
            "allocations": len(alloc_raw), "restrictions": len(rules_raw),
            "deals": len(pipe), "plans": 0}
    prog = ob.progress(have)

    st.subheader("Set up your estate")
    st.markdown(ob.next_step(have))
    st.progress(prog["required_done"] / max(prog["required_total"], 1))
    st.caption(f"{prog['required_done']} of {prog['required_total']} required inputs loaded. "
               "Turn off **Use sample data** in the sidebar to upload your files.")
    st.dataframe(ob.checklist(have), width="stretch", hide_index=True)
    st.info("Two files get you a working plan: the **floor inventory** and the "
            "**planning workbook**. The other four add capability and can follow later.")
    with st.expander("What each file needs to contain"):
        for spec in ob.INPUTS:
            if not spec["columns"]:
                continue
            st.markdown(f"**{spec['name']}** — "
                        + ("required" if spec["required"] else "optional")
                        + "  \n`" + ", ".join(spec["columns"]) + "`"
                        + (f"  \noptional: `{', '.join(spec['optional_columns'])}`"
                           if spec["optional_columns"] else ""))
    st.stop()

problems = es.validate_floors(floors_raw)
if problems:
    st.error("Floor inventory cannot be used:\n\n" + "\n".join(f"- {p}" for p in problems))
    st.stop()
floors = es.prepare(floors_raw)

alloc_problems = mv.validate_allocations(alloc_raw, floors)
alloc = None if alloc_problems else mv.prepare_moves(alloc_raw, floors)

rule_problems = rx.validate(rules_raw)
if rule_problems:
    st.warning("Restrictions file ignored:\n\n" + "\n".join(f"- {p}" for p in rule_problems))
    rules_raw = rx.empty()
if "answered_rules" not in st.session_state:
    st.session_state["answered_rules"] = rx.empty()
@st.cache_data(show_spinner="Reading the planning workbook…")
def load_os_cached(data):
    return os1.read_workbook(io.BytesIO(data))


try:
    os_src = load_os_cached(open("data/sample_one_source.xlsx", "rb").read())
except Exception:
    os_src = None      # nothing bundled — the Data quality tab asks for a file

RULES = rx.Rules(pd.concat([rules_raw, st.session_state["answered_rules"]], ignore_index=True))

unknown = demand.attrs.get("unrecognised_metrics") or []
if unknown:
    st.warning("Metric rows not recognised and ignored: " + ", ".join(map(str, unknown)))

with st.expander("⬇ Blank templates — the columns each file needs", expanded=False):
    st.caption("Empty files with the right headers. Fill them in and upload from the sidebar.")
    st.download_button("📦 Download all templates", sample_bundle(),
                       file_name="floorcast_templates.zip", mime="application/zip",
                       key="dl_bundle_main")
    tcols = st.columns(3)
    for i, (name, tag, fname, kind, cols, desc) in enumerate(TEMPLATES):
        c = tcols[i % 3]
        c.markdown(f"**{name}**  \n<span style='color:#5B6B6B;font-size:0.82em'>{tag} — {desc}"
                   "</span>", unsafe_allow_html=True)
        c.download_button("Download", _template_bytes(kind, cols), file_name=fname,
                          mime="text/csv" if kind == "csv" else
                          "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                          width="stretch", key="dlm_" + fname)

(tab_role, tab_what, tab_estate, tab_one, tab_demand, tab_ramp, tab_map,
 tab_scen, tab_qa) = st.tabs(
    ["👥 My view", "🎛 What if", "🏢 Estate", "📑 One source", "📈 Demand",
     "🎯 Expansion plan", "🗺 Floor map", "🔮 Scenarios", "🔍 Data quality"])



# ═══════════════════════════════════════════════ 0 · MY VIEW
with tab_role:
    who = st.radio("I work in", list(rl.ROLES),
                   horizontal=True, key="role_pick",
                   format_func=lambda r: f"{rl.ROLES[r]['icon']} {r}")
    meta = rl.ROLES[who]
    st.markdown(f"#### {meta['question']}")
    st.caption(f"**What this view holds:** {meta['cares']}  \n"
               f"**The step you take:** {meta['action']}")
    st.divider()

    r_period = st.select_slider("Period", options=dr.week_options(demand),
                                value=dr.peak_week(demand), key="role_period")
    r_res = sc.compute(floors, demand, pipe, sc.levers(period=r_period))
    r_tot = r_res["totals"]

    if who == "Leadership":
        lvl_l = "geo" if "geo" in floors.columns else "site"
        rt = th.rag(floors, lvl_l)
        counts = rt["status"].value_counts().to_dict() if not rt.empty else {}
        sc_ = st.columns(3)
        for c, name in zip(sc_, ["Critical", "Strained", "Healthy"]):
            c.metric(name, counts.get(name, 0),
                     delta_color="inverse" if name != "Healthy" else "normal")
        for a in th.alerts(rt, lvl_l)[:3]:
            (st.error if "critical" in a else st.warning)(a)
        if not rt.empty:
            st.dataframe(_status_style(
                rt[[lvl_l, "total_seats", "occupancy_%", "critical_above", "status"]]),
                width="stretch", hide_index=True)
        st.caption("Regions are rated against a limit that scales with their size — a small "
                   "estate goes red sooner because it has less room to flex.")

        st.markdown("##### The estate in numbers")
        vals = rl.leadership(floors, {**es.estate_totals(floors), **{}})
        cols = st.columns(len(vals))
        for c, (k, v) in zip(cols, vals.items()):
            c.metric(k, v)
        st.markdown("##### Where the money is stuck")
        tb = es.trapped_breakdown(floors)
        st.dataframe(tb[["trapped_reason", "definition", "seats", "share_%", "releasable"]],
                     width="stretch", hide_index=True)
        rec = int(tb.loc[tb["releasable"], "seats"].sum())
        st.info(f"**{rec:,} seats** sit against reasons that could be recovered without new "
                f"space — {rec / max(es.estate_totals(floors)['total_seats'], 1) * 100:.0f}% of "
                "the estate. That is the cheapest capacity available, and the first question "
                "to ask before approving a lease.")

    elif who == "Operations":
        k = st.columns(3)
        k[0].metric("Seats needed", f"{r_tot['incremental_need']:,}")
        k[1].metric("Seats usable", f"{r_tot['usable']:,}")
        k[2].metric("Short", f"{r_tot['shortfall']:,}", delta_color="inverse",
                    delta=f"{r_tot['sites_short']} site(s)" if r_tot['sites_short'] else None)
        st.dataframe(rl.operations(r_res["sites"]), width="stretch", hide_index=True)
        if r_tot["shortfall"]:
            st.error(f"Escalate now for {r_tot['sites_short_names']} — a ramp that is short at "
                     "plan stage is far cheaper to fix than one short at go-live.")
        else:
            st.success("Every site lands. This is the evidence to commit the ramp on.")

    elif who == "Facility":
        fac = rl.facility(floors)
        k = st.columns(3)
        k[0].metric("Renovation in pipeline", f"{int(floors['expansion_space'].sum()):,} seats")
        k[1].metric("Floors with idle space", int((fac.get('idle_%', pd.Series(dtype=float)) > 25).sum()))
        k[2].metric("Unusable seats", f"{int(floors['trapped'].sum()):,}")
        st.dataframe(fac, width="stretch", hide_index=True)
        st.caption("Sorted by what a renovation would unlock. The lead time column is the one "
                   "that decides whether it helps this ramp or the next one.")

        st.markdown("##### Coming up in the next six months")
        st.caption("Work already committed, in date order. This is the queue, not a wish list.")
        sched_f = ex.expansion_schedule(floors)
        if sched_f.empty:
            st.info("No dated renovation in the floor inventory.")
        else:
            soon = sched_f[sched_f.get("weeks_away", 99) <= 26] if "weeks_away" in sched_f \
                else sched_f
            st.dataframe(soon, width="stretch", hide_index=True)
            st.info(f"**{int(soon['expansion_space'].sum()):,} seats** are due to land in the "
                    f"next six months across {len(soon)} project(s).")

        st.markdown("##### Space held but not used")
        if alloc is None or alloc.empty:
            st.info("Add the allocations file to see under-used floors.")
        else:
            oc_f = cm.over_contracted(floors, alloc_raw)
            if oc_f.empty:
                st.success("No floor is held by one programme and left largely empty.")
            else:
                st.dataframe(oc_f, width="stretch", hide_index=True)
                st.info(cm.over_contracted_note(oc_f)
                        + " Reallocating them needs an account conversation, not a build.")

    elif who == "Security":
        st.metric("Seats stranded by segregation",
                  f"{int(floors.loc[floors['trapped_reason'].eq('segregation'), 'trapped'].sum()):,}")
        st.markdown("##### Rules in force")
        rs = RULES.summary()
        st.dataframe(rs if not rs.empty else pd.DataFrame({"note": ["No restrictions loaded"]}),
                     width="stretch", hide_index=True)
        st.markdown("##### Enclosed or shared")
        st.caption("The request that starts a programme says whether it can share a floor. "
                   "Set it here and the plan is checked against it.")
        if alloc is None or alloc.empty:
            st.info("Add the allocations file to check space requirements.")
        else:
            accts_s = sorted(alloc["account"].unique())
            need_map = {}
            cols_s = st.columns(min(len(accts_s), 4) or 1)
            for i, a_ in enumerate(accts_s):
                need_map[a_] = cols_s[i % len(cols_s)].selectbox(
                    a_, list(cm.SPACE_TYPES), key=f"space_{a_}",
                    format_func=lambda k: k.capitalize())
            chk = cm.space_check(alloc_raw, need_map)
            if not chk.empty:
                st.dataframe(chk, width="stretch", hide_index=True)
                bad = chk[chk["meets_requirement"] == "no"]
                if not bad.empty:
                    st.error(", ".join(f"**{r['account']}** needs {r['needs']} space but "
                                       f"shares {int(r['shared_floors'])} floor(s)"
                                       for _, r in bad.iterrows())
                             + ". Either move them, or partition the floor.")
                else:
                    st.success("Every programme is in the kind of space its request asked for.")

        st.info("Zone-level segregation is checked on the Floor map tab for any floor with a "
                "plan loaded. A plan that fits on capacity can still breach a segregation "
                "clause — the two are separate tests.")

    elif who == "WFM":
        if os_src is None or os_src["long"].empty:
            st.info("Load the planning workbook on the Data quality tab to see ratio behaviour.")
        else:
            w = rl.wfm(os_src["long"])
            if not w:
                st.info("No usable seat ratios found in the workbook.")
            else:
                k = st.columns(2)
                k[0].metric("Programmes with a ratio", f"{int(w['by_geo']['programmes'].sum()):,}")
                k[1].metric("Ratios above the ceiling", len(w["above_ceiling"]))
                st.markdown("##### Ratios that cannot be right")
                st.dataframe(w["above_ceiling"], width="stretch", hide_index=True)
                st.caption("A ratio is people per seat. Anything above about four is worth "
                           "treating as a data entry error rather than a plan.")
                st.markdown("##### Ratio by geo")
                st.dataframe(w["by_geo"], width="stretch", hide_index=True)

    elif who == "IT":
        q = rl.it(floors)
        st.metric("Seats waiting on provisioning", f"{q['seats']:,}")
        st.caption("Seats built but not usable until IT provisions them, and the network "
                   "work that a renovation will create.")
        if q["queue"].empty:
            st.success("Nothing waiting on IT.")
        else:
            st.dataframe(q["queue"], width="stretch", hide_index=True)
            st.info("These seats exist and are empty for want of a build. Provisioning them is "
                    "usually the cheapest capacity in the estate — no lease, no construction.")

        st.markdown("##### Network work coming in the next six months")
        sched_it = ex.expansion_schedule(floors)
        if not sched_it.empty:
            soon_it = sched_it[sched_it.get("weeks_away", 99) <= 26] if "weeks_away" in sched_it \
                else sched_it
            if not soon_it.empty:
                st.dataframe(soon_it[[c for c in ("site", "floor", "expansion_space",
                                                  "expansion_expected", "lands")
                                      if c in soon_it.columns]],
                             width="stretch", hide_index=True)
                st.caption(cm.move_note(int(soon_it["expansion_space"].sum()))
                           + " Each of these floors needs the network in place before the "
                             "seats are usable.")

    elif who == "Client":
        if alloc is None or alloc.empty:
            st.info("Add the allocations file to see a client footprint.")
        else:
            accounts_all = sorted(alloc["account"].unique())
            vc1, vc2 = st.columns([1, 2])
            acct = vc1.selectbox("Viewing as", accounts_all, key="role_acct")
            vc2.caption("**This is a preview of what that client sees when they sign in.** "
                        "A client account is locked to its own footprint — the picker exists "
                        "here so you can check each one. Their view is cut from the data, not "
                        "hidden in the page: everything below is filtered to this client "
                        "before it reaches the screen.")

            c = rl.client_safe_view(alloc, floors, acct, demand=demand, period=r_period)
            if not c:
                st.info(f"{acct} holds no seats in the allocations file.")
            else:
                k = st.columns(4)
                k[0].metric("Seats held", f"{c['seats']:,}")
                k[1].metric("Seats needed", f"{c['need']:,}" if c.get("need") is not None else "—")
                k[2].metric("Free beside them", f"{c['headroom']:,}")
                k[3].metric("Floors", c["floors"])

                if c.get("verdict"):
                    if c.get("growth") is not None and c["growth"] > c["headroom"]:
                        st.error(c["verdict"])
                    elif c.get("growth", 0) > 0:
                        st.success(c["verdict"])
                    else:
                        st.info(c["verdict"])

                st.markdown("##### Where you sit")
                st.dataframe(c["detail"], width="stretch", hide_index=True)

                if c["shares_a_floor"]:
                    st.warning("Some of your seats are on floors shared with another client. "
                               "Raise it with your account team if your contract requires "
                               "dedicated space.")
                else:
                    st.success("All of your seats are on floors held by no other client.")

                with st.expander("What this view deliberately does not show"):
                    st.markdown(
                        "- Which other clients are in the building, or where  \n"
                        "- How full the wider estate is, or what is free elsewhere  \n"
                        "- Any other client's headcount, growth or contract  \n\n"
                        "A client sharing a floor is told that they share it — that is their "
                        "contract — but not with whom. The internal Security and Operations "
                        "views hold the full picture."
                    )

    elif who == "Finance":
        oc_fin = cm.over_contracted(floors, alloc_raw) if alloc is not None \
            and not alloc.empty else pd.DataFrame()
        fin = rl.finance(floors, alloc_raw, es.estate_totals(floors), oc_fin)
        show = {k: v for k, v in fin.items() if not k.startswith("_")}
        fc_ = st.columns(len(show))
        for c, (k, v) in zip(fc_, show.items()):
            c.metric(k, v)
        st.error(f"**{fin['_recoverable']:,} seats** — {fin['_recoverable_pct']}% of the "
                 "estate — are paid for and returning nothing: unusable seats plus floors "
                 "held but barely used.")

        st.markdown("##### Seats paid for against seats used")
        bp = cm.billing_position(floors, alloc_raw)
        if "note" in bp.columns:
            st.info("No contracted-seat figures supplied. Add a `seats_contracted` column to "
                    "the allocations file and this becomes paid against used, by floor.")
        st.dataframe(bp.head(20), width="stretch", hide_index=True)

        if not oc_fin.empty:
            st.markdown("##### Space held but barely used")
            st.dataframe(oc_fin, width="stretch", hide_index=True)
            st.caption("Recovering these needs an account conversation, not a build — which "
                       "makes them the cheapest seats on the table.")

        st.markdown("##### Revenue per seat")
        if rates_raw is None or rates_raw.empty:
            st.info("Add a rate card — site, rate_per_seat — and this becomes revenue per "
                    "seat, the cost of idle space, and what the unusable seats would be "
                    "worth if they could be sold.")
        else:
            rp = rv.position(floors, alloc_raw, rates_raw)
            tot_r = rv.totals(rp)
            rk = st.columns(4)
            rk[0].metric("Revenue per seat sold", f"{tot_r['currency']}{tot_r['per_seat_sold']:,.0f}")
            rk[1].metric("Revenue per seat owned",
                         f"{tot_r['currency']}{tot_r['per_seat_owned']:,.0f}",
                         delta=f"-{tot_r['per_seat_sold'] - tot_r['per_seat_owned']:,.0f} diluted",
                         delta_color="inverse")
            rk[2].metric("Seats earning nothing", f"{tot_r['idle_seats']:,}")
            rk[3].metric("Cost of that idle space",
                         f"{tot_r['currency']}{tot_r['cost_of_idle']:,.0f}")
            st.caption("**Per seat sold** is the rate. **Per seat owned** spreads it across "
                       "every seat in the building, including the ones earning nothing. The "
                       "gap between them is what empty and unusable space costs.")
            st.dataframe(rp[["site", "total_seats", "seats_sold", "rate", "revenue",
                             "rev_per_seat_sold", "rev_per_seat_owned", "dilution",
                             "cost_of_idle"]], width="stretch", hide_index=True)

            rec_v = rv.recovery_value(floors, rates_raw,
                                      ["segregation", "layout", "it_not_ready",
                                       "under_renovation"])
            if rec_v.get("seats"):
                st.success(f"The {rec_v['seats']:,} recoverable unusable seats would be worth "
                           f"**{tot_r['currency']}{rec_v['value']:,.0f}** if they could be "
                           "sold. That is the trapped-seat case in the currency that gets it "
                           "approved.")

            st.markdown("##### Revenue per seat across the horizon")
            fc_r = rv.forecast(demand, floors, rates_raw)
            if not fc_r.empty:
                st.line_chart(fc_r.set_index("period")[["rev_per_seat_sold",
                                                        "rev_per_seat_owned"]], height=240)
                st.caption("A rising owned line means seats filling up. A falling one means "
                           "the estate growing faster than the demand for it.")
                exp_seats = int(floors["expansion_space"].sum()) \
                    if "expansion_space" in floors.columns else 0
                if exp_seats:
                    eff = rv.expansion_effect(fc_r, exp_seats)
                    (st.warning if eff["change"] < 0 else st.success)(
                        f"Taking the {exp_seats:,} seats of planned expansion with no extra "
                        f"revenue moves revenue per seat owned from "
                        f"{tot_r['currency']}{eff['before']:,.0f} to "
                        f"{tot_r['currency']}{eff['after']:,.0f}. {eff['verdict']}.")

        st.markdown("##### Renovation spend in the plan")
        yr_fin = ex.expansion_by_year(floors)
        if yr_fin.empty:
            st.info("No dated renovation to profile.")
        else:
            st.dataframe(yr_fin, width="stretch", hide_index=True)

    else:  # PMO
        if os_src is None or os_src["programmes"].empty:
            st.info("Load the planning workbook on the Data quality tab to see collection status.")
        else:
            pr = os_src["programmes"]
            iss = os1.quality_report(os_src["long"])
            k = st.columns(4)
            done = int(pr["status"].eq("Updated").sum())
            k[0].metric("Programmes", f"{len(pr):,}")
            k[1].metric("Collected", f"{done / max(len(pr), 1) * 100:.0f}%")
            k[2].metric("Outstanding", f"{len(pr) - done:,}")
            k[3].metric("Flagged to check", f"{iss['programme'].nunique() if not iss.empty else 0:,}")
            st.markdown("##### Who to chase")
            ch = os1.chase_list(pr)
            st.dataframe(ch.head(25), width="stretch", hide_index=True)
            st.download_button("📋 Chase list (CSV)", ch.to_csv(index=False).encode(),
                               file_name="not_updated.csv", mime="text/csv", key="pmo_chase")

# ═══════════════════════════════════════════════ 0 · WHAT IF
with tab_what:
    st.caption("**Move the levers and watch the answer change.** The files are loaded — nothing "
               "needs uploading again. Save a scenario to put it beside another one.")

    periods_w = dr.week_options(demand)
    pk_w = dr.peak_week(demand)
    if "saved_scenarios" not in st.session_state:
        st.session_state["saved_scenarios"] = {}

    lev, res = st.columns([1, 2.35], gap="large")

    with lev:
        st.markdown("##### Levers")
        w_period = st.select_slider("Period", options=periods_w,
                                    value=pk_w if pk_w in periods_w else periods_w[-1],
                                    key="w_period")
        w_uplift = st.slider("Demand vs plan", -0.25, 0.50, 0.0, 0.05,
                             format="%+.0f%%", key="w_uplift",
                             help="What if volumes land above or below forecast")
        st.markdown("**Trapped seats**")
        tb_w = es.trapped_breakdown(floors)
        rec = tb_w.loc[tb_w["releasable"], "trapped_reason"].tolist()
        w_reasons = st.multiselect("Assume recovered", tb_w["trapped_reason"].tolist(),
                                   default=rec, key="w_reasons",
                                   label_visibility="collapsed")
        w_frac = st.slider("How much comes back", 0.0, 1.0, 0.0, 0.05,
                           format="%.0f%%", key="w_frac")
        st.markdown("**Renovation**")
        w_exp = st.toggle("Count expansion space", value=True, key="w_exp")
        w_hor = st.slider("Landing within (weeks)", 0, 52, 12, 2, key="w_hor",
                          disabled=not w_exp)
        st.markdown("**Nesting**")
        w_nest = st.slider("Seats held for new starters", 0.0, 0.5, 0.0, 0.05,
                           format="%.0f%%", key="w_nest",
                           help="A share of each period's growth, held while new starters "
                                "finish nesting before they reach production.")
        st.markdown("**Deals not yet won**")
        w_pipe = st.radio("Deals not yet won", ["exclude", "weighted", "full"], index=0,
                          key="w_pipe", horizontal=True,
                          format_func=lambda m: {"exclude": "Ignore them",
                                                 "weighted": "Likely ones",
                                                 "full": "All of them"}[m],
                          label_visibility="collapsed")

    demand_w = demand
    nest_total = 0
    if w_nest > 0:
        nn = cm.nesting_need(demand, w_period, nest_pct=w_nest)
        if not nn.empty:
            nest_total = int(nn["nesting"].sum())
            add = nn.rename(columns={"Account": "Account", "Site": "Site"})[
                ["Account", "Site", "nesting"]]
            demand_w = demand.merge(add, on=["Account", "Site"], how="left")
            demand_w["nesting"] = demand_w["nesting"].fillna(0)
            mask = demand_w["week"] == w_period
            demand_w.loc[mask, "seats"] = (demand_w.loc[mask, "seats"]
                                           + demand_w.loc[mask, "nesting"]).astype(int)

    lv = sc.levers(period=w_period, demand_uplift=w_uplift, release_reasons=w_reasons,
                   release_fraction=w_frac, include_expansion=w_exp, horizon_weeks=w_hor,
                   pipeline_mode=w_pipe)
    now = sc.compute(floors, demand_w, pipe, lv)
    base_lv = sc.levers(period=w_period)
    base = sc.compute(floors, demand, pipe, base_lv)
    if nest_total:
        st.caption(f"Nesting adds {nest_total:,} seats to this period's requirement.")
    T, BT = now["totals"], base["totals"]

    with res:
        k = st.columns(4)
        k[0].metric("Seats needed", f"{T['incremental_need']:,}",
                    delta=f"{T['incremental_need'] - BT['incremental_need']:+,}" if T['incremental_need'] != BT['incremental_need'] else None,
                    delta_color="inverse")
        k[1].metric("Usable seats", f"{T['usable']:,}",
                    delta=f"{T['usable'] - BT['usable']:+,}" if T['usable'] != BT['usable'] else None)
        k[2].metric("Shortfall", f"{T['shortfall']:,}",
                    delta=f"{T['shortfall'] - BT['shortfall']:+,}" if T['shortfall'] != BT['shortfall'] else None,
                    delta_color="inverse")
        k[3].metric("Sites short", T["sites_short"],
                    delta=f"{T['sites_short'] - BT['sites_short']:+d}" if T['sites_short'] != BT['sites_short'] else None,
                    delta_color="inverse")

        if T["shortfall"] == 0:
            st.success(f"**Everything fits.** {T['usable']:,} usable seats against "
                       f"{T['incremental_need']:,} needed.")
        else:
            st.error(f"**{T['shortfall']:,} seats short** at {T['sites_short_names']}. "
                     f"{T['trapped_left']:,} trapped seats are still not counted.")

        wf = sc.waterfall(now)
        st.markdown("###### How the scenario gets there")
        st.bar_chart(wf.set_index("step")["seats"], height=210)

        st.markdown("###### Site by site")
        show = now["sites"][["site", "allocated", "incremental_need", "available",
                             "released", "expansion_counted", "usable", "gap", "status"]]
        st.dataframe(show, width="stretch", hide_index=True)

        st.caption(f"This scenario: {sc.describe(lv)}")
        sv1, sv2, sv3 = st.columns([2, 1, 1])
        nm = sv1.text_input("Name this scenario", value="", placeholder="e.g. Release half, no renovation",
                            label_visibility="collapsed")
        if sv2.button("Save scenario", width="stretch"):
            st.session_state["saved_scenarios"][nm or f"Scenario {len(st.session_state['saved_scenarios']) + 1}"] = now
            st.rerun()
        if sv3.button("Clear saved", width="stretch"):
            st.session_state["saved_scenarios"] = {}
            st.rerun()

        # ── the floor, under this scenario
        st.markdown("###### The floor, under this scenario")
        try:
            wf_plan = load_plan(plan_file.getvalue(), plan_file.name) if plan_file \
                else load_plan(open("data/sample_seat_inventory.csv", "rb").read(),
                               "sample_seat_inventory.csv")
        except Exception:
            wf_plan = None

        if wf_plan is None or wf_plan["seats"].empty:
            st.info("Load a floor plan to see the map respond to these levers.")
        else:
            wseats = wf_plan["seats"].copy()
            if "zone_type" not in wseats.columns:
                wseats["zone_type"] = "Production"
            for cn, vv in [("country", "-"), ("city", "-"), ("site", "SITE"),
                           ("building", "B1"), ("tower", "T1"), ("floor", "F1")]:
                if cn not in wseats.columns:
                    wseats[cn] = vv
            allocz = sorted(wseats.loc[wseats["zone_type"].astype(str).str.lower()
                                       == "production", "zone"].unique())
            # the release lever frees trapped seats on the plan too, so the map
            # answers the same question as the numbers above it
            if "seat_status" in wseats.columns and w_reasons and w_frac > 0:
                cand = wseats.index[wseats["seat_status"].eq("trapped")
                                    & wseats["trapped_reason"].isin(w_reasons)]
                take = int(round(len(cand) * w_frac))
                if take:
                    wseats.loc[cand[:take], "seat_status"] = "usable"
                    wseats.loc[cand[:take], "trapped_reason"] = ""
            still_trapped = int(wseats.get("seat_status",
                                           pd.Series(dtype=str)).eq("trapped").sum())
            cap_w = int((wseats["zone"].isin(allocz)
                         & wseats.get("seat_status",
                                      pd.Series("usable", index=wseats.index))
                         .ne("trapped")).sum())
            labs_w, sites_w = dr.site_options(demand)
            fitw = {lb: abs(int(dr.slice_week(demand, w_period, sn)["seats"].sum()) - cap_w)
                    for lb, sn in zip(labs_w, sites_w)}
            best_w = min(fitw, key=fitw.get) if fitw else None
            mc1, mc2 = st.columns([1, 1])
            wsite_lab = mc1.selectbox("Floor holds demand for", labs_w,
                                      index=labs_w.index(best_w) if best_w in labs_w else 0,
                                      key="wf_site")
            wlevel = mc2.radio("Colour by", ["Account", "LOB"], horizontal=True, key="wf_lvl")
            wsite = sites_w[labs_w.index(wsite_lab)]

            wsl = dr.slice_week(demand, w_period, wsite)
            if wsl.empty:
                st.info("No demand for that site in this period.")
            else:
                wsl = wsl.copy()
                wsl["seats"] = np.ceil(wsl["seats"] * (1 + w_uplift)).astype(int)
                if w_pipe != "exclude" and pipe is not None and not pipe.empty:
                    pw = dr.pipeline_seats(pipe, w_pipe, 1.2)
                    pw = pw[pw["site"] == wsite]
                    for _, prow in pw.iterrows():
                        if int(prow["seats"]) > 0:
                            wsl = pd.concat([wsl, pd.DataFrame([{
                                "account": f"{prow['account']} (not yet won)",
                                "lob": "New business", "seats": int(prow["seats"])}])],
                                ignore_index=True)
                wsl["site"] = wseats["site"].iloc[0]
                wsl["building"] = wseats["building"].iloc[0]
                wsl["tower"] = wseats["tower"].iloc[0]
                wsl["floor"] = None
                wassigned, _, wunplaced = sm.allocate_seats(wsl, wseats,
                                                            allocatable_zones=allocz)
                placed = int(wassigned["account"].notna().sum())
                mm = st.columns(4)
                mm[0].metric("Seats wanted here", f"{int(wsl['seats'].sum()):,}")
                mm[1].metric("Usable on this floor", f"{cap_w:,}")
                mm[2].metric("Still unusable", f"{still_trapped:,}",
                             delta_color="inverse")
                mm[3].metric("Left empty", f"{cap_w - placed:,}")
                st.plotly_chart(
                    fr.plotly_map(wassigned, background=wf_plan.get("background"),
                                  extent=wf_plan.get("extent"),
                                  level="account" if wlevel == "Account" else "lob"),
                    width="stretch", key="wf_map")
                st.caption("Move a lever above and the colours here follow. Gold outlines are seats "
                           "that exist but cannot be used — releasing them above brings them "
                           "into play here. Hover any seat for its id, zone and holder.")
                if not wunplaced.empty:
                    st.warning(f"{int(wunplaced['short_by'].sum()):,} seat(s) wanted here have "
                               "nowhere to sit on this floor under this scenario.")

    saved = st.session_state["saved_scenarios"]
    if saved:
        st.markdown("##### Saved scenarios")
        st.dataframe(sc.compare({"Baseline": base, **saved}), width="stretch", hide_index=True)
        st.caption("Baseline is the same period with no levers applied. "
                   "**vs baseline** is seats of shortfall removed — higher is better.")

# ═══════════════════════════════════════════════ 1 · ESTATE
with tab_estate:
    st.caption("**How much space do we have, and how much of it can we actually use?**")

    rag_level = "geo" if "geo" in floors.columns else "site"
    rag_tbl = th.rag(floors, rag_level)
    for a in th.alerts(rag_tbl, rag_level)[:4]:
        (st.error if "critical" in a else st.warning)(a)
    if not rag_tbl.empty and rag_tbl["status"].eq("Healthy").all():
        st.success("Every region is inside its occupancy limit.")

    t = es.estate_totals(floors)
    c = st.columns(6)
    c[0].metric("Total seats", f"{t['total_seats']:,}")
    c[1].metric("Allocated", f"{t['allocated']:,}")
    c[2].metric("Available", f"{t['available']:,}")
    c[3].metric("Trapped", f"{t['trapped']:,}", delta=f"{t['trapped_%']}% of estate",
                delta_color="inverse")
    c[4].metric("Occupancy", f"{t['occupancy_%']}%")
    c[5].metric("Expansion space", f"{t['expansion_space']:,}",
                help="Not built yet — unlocked by renovation")

    lvl = st.radio("Roll up by", ["geo", "country", "city", "site", "floor_id"],
                   horizontal=True, index=0)
    st.dataframe(es.rollup(floors, lvl), width="stretch", hide_index=True)

    st.markdown("#### Occupancy and utilisation")
    st.caption("**Occupancy** is how much of the building has been handed out. "
               "**Utilisation** is how much of what was handed out is genuinely needed. "
               "A site can be fully occupied and poorly utilised — every seat assigned, "
               "half of them not required — and that is the case worth finding.")
    util_period = st.select_slider("Period", options=dr.week_options(demand),
                                   value=dr.peak_week(demand), key="est_util")
    ru = es.with_utilisation(es.rollup(floors, "site"), demand, util_period, "site")
    if "utilisation_%" in ru.columns:
        st.dataframe(ru[["site", "total_seats", "allocated", "occupancy_%", "required",
                         "utilisation_%", "verdict"]], width="stretch", hide_index=True)
        low = ru[ru["utilisation_%"] < 80]
        if not low.empty:
            st.warning(", ".join(f"**{r['site']}** holds {int(r['allocated'])} seats but needs "
                                 f"{int(r['required'])}" for _, r in low.head(3).iterrows())
                       + " — space handed out and not required.")

    st.markdown("#### Occupancy against the limit")
    st.caption("A large estate can run hot because there is somewhere to flex to; a small one "
               "cannot. The limit therefore scales with the size of the region rather than "
               "being one number for everywhere.")
    rl_lvl = lvl if lvl in ("geo", "country", "city", "site") else "site"
    rag_show = th.rag(floors, rl_lvl)[[rl_lvl, "total_seats", "allocated", "occupancy_%",
                                       "strained_above", "critical_above", "status",
                                       "headroom_seats"]]
    st.dataframe(_status_style(rag_show), width="stretch", hide_index=True)

    st.markdown("#### Trapped seats")
    st.caption("Seats that physically exist and are paid for, but cannot be sold to a "
               "client — stranded inside another client's secure zone, waiting on IT, or "
               "in a block too small to allocate. Tracked by reason, because the reason "
               "decides whether it can be recovered.")
    tb = es.trapped_breakdown(floors)
    st.dataframe(tb, width="stretch", hide_index=True)
    st.markdown("##### Against what the business will carry")
    st.caption("Unusable seats have to be justified, not just reported — there is a limit on "
               "how many the organisation will hold, and the monthly report has to defend them.")
    tp = th.trapped_position(floors, rag_level)
    st.dataframe(tp, width="stretch", hide_index=True)
    over = tp[tp["within_budget"] == "no"]
    if not over.empty:
        st.warning(", ".join(f"{r[rag_level]} is {int(r['over_budget'])} seats over budget"
                             for _, r in over.iterrows())
                   + " — each of those needs a reason leadership will accept.")

    st.markdown("#### Floors held but barely used")
    st.caption("A programme holding a whole floor while occupying part of it is the cheapest "
               "capacity there is — the seats exist, they are already paid for, and nothing "
               "has to be built.")
    if alloc is None or alloc.empty:
        st.info("Add the allocations file to see which floors are held but under-used.")
    else:
        oc = cm.over_contracted(floors, alloc_raw)
        if oc.empty:
            st.success("No floor is held by a single programme that is using well under it.")
        else:
            st.dataframe(oc, width="stretch", hide_index=True)
            st.info(cm.over_contracted_note(oc))

    st.markdown("##### The defence for each one")
    st.caption("A reason with a route out is defensible. One without is a seat the business "
               "is simply carrying, and that is the conversation leadership will open.")
    st.dataframe(th.justification(floors, rag_level), width="stretch", hide_index=True)

    rel = int(tb.loc[tb["releasable"], "seats"].sum())
    st.info(f"**{rel:,} of {t['trapped']:,} trapped seats** sit against reasons that a "
            f"partition move, an IT rollout or a relocation could address — "
            f"{rel / t['total_seats'] * 100:.1f}% of the estate. "
            "Model that in the Scenarios tab before committing to new space.")


# ═══════════════════════════════════════════════ 3 · ONE SOURCE
with tab_one:
    st.caption("**Everything in one table.** Filter down to a geography and see the lot — "
               "floors, who is on them, what is needed, what is unusable and why — without "
               "moving between views.")

    ov_period = st.select_slider("Period for the demand columns",
                                 options=dr.week_options(demand),
                                 value=dr.peak_week(demand), key="ov_period")
    table = ov.build(floors, alloc_raw, demand, ov_period)

    fcols = [c for c in ("geo", "country", "city", "site", "trapped_reason",
                         "expansion_status") if c in table.columns]
    sel = {}
    fc = st.columns(len(fcols))
    for col, name in zip(fc, fcols):
        opts = ov.filter_options(table, name, sel)
        sel[name] = col.multiselect(name.replace("_", " "), opts, key="ovf_" + name)
    view = ov.apply_filters(table, sel)

    s_ = ov.summarise(view)
    if s_:
        k = st.columns(6)
        k[0].metric("Floors", s_["floors"])
        k[1].metric("Sites", s_["sites"])
        k[2].metric("Seats", f"{s_['total_seats']:,}")
        k[3].metric("Occupancy", f"{s_['occupancy_%']}%")
        k[4].metric("Unusable", f"{s_['trapped']:,}")
        k[5].metric("Coming", f"{s_['expansion_space']:,}")

    groups = st.multiselect("Columns", list(ov.COLUMN_GROUPS),
                            default=["Where", "Capacity", "Unusable"], key="ov_groups")
    show = [c for g in groups for c in ov.COLUMN_GROUPS[g] if c in view.columns]
    st.dataframe(view[show] if show else view, width="stretch", hide_index=True, height=430)
    st.caption(f"{len(view)} of {len(table)} floors shown. Filters cascade, so each one only "
               "offers what is still available under the others.")
    st.download_button("📊 This view (Excel)", _xlsx({"One source": view[show] if show else view}),
                       file_name="Floorcast_OneSource.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ═══════════════════════════════════════════════ 2 · DEMAND
with tab_demand:
    st.caption("**How many seats does each client need, and when?**")
    periods = dr.week_options(demand)
    labels, sites = dr.site_options(demand)
    cc = st.columns(2)
    site_pick = cc[0].selectbox("Site", ["All sites"] + labels, key="dem_site")
    site = None if site_pick == "All sites" else sites[labels.index(site_pick)]
    pk = dr.peak_week(demand, site)
    period = cc[1].select_slider("Period", options=periods,
                                 value=pk if pk in periods else periods[-1])
    if period == pk:
        cc[1].caption("Peak period for this selection.")

    d = demand if site is None else demand[demand["Site"] == site]
    m = st.columns(4)
    cur = d[d["week"] == period]
    m[0].metric("Seats required", f"{int(cur['seats'].sum()):,}")
    if cur["hc"].notna().any():
        m[1].metric("Headcount", f"{int(cur['hc'].sum()):,}")
    if cur["seat_ratio"].notna().any():
        m[2].metric("Avg seat ratio", f"{cur['seat_ratio'].mean():.2f}")
    m[3].metric("Periods in file", len(periods))

    piv = d.pivot_table(index="week", columns="Account", values="seats",
                        aggfunc="sum").fillna(0)
    st.line_chart(piv)
    st.caption("Seats required per period by account.")
    st.dataframe(cur, width="stretch", hide_index=True)

# ═══════════════════════════════════════════════ 3 · EXPANSION PLAN
with tab_ramp:
    st.caption("**Does the expansion fit, and if not who has to do something?** New space is "
               f"approved once a site is demonstrably using what it holds — below "
               f"{ex.OCCUPANCY_GATE:.0f}% occupancy the answer is fill that first.")

    e1, e2, e3 = st.columns(3)
    eperiod = e1.select_slider("Period", options=dr.week_options(demand),
                               value=dr.peak_week(demand), key="exp_period")
    count_sunset = e2.toggle("Count seats coming back", value=True,
                             help="Programmes winding down return seats. Netting them off "
                                  "avoids asking for space that is about to free up.")
    view_mode = e3.radio("View", ["By site", "By year"], horizontal=True, key="exp_view")

    need_by_site = dr.need_by_site(demand, eperiod)
    pos = ex.site_position(floors, need_by_site)
    if count_sunset and not sunset_raw.empty:
        pos = ex.net_position(pos, sunset_raw)

    gs = ex.gate_summary(pos)
    esc = ex.escalations(pos, floors)
    k = st.columns(4)
    k[0].metric("Sites asking for space", gs.get("sites_asking", 0))
    k[1].metric("Blocked by the gate", gs.get("blocked_by_gate", 0), delta_color="inverse")
    k[2].metric("Seats behind the gate", f"{gs.get('seats_blocked', 0):,}")
    k[3].metric("Escalations to raise", len(esc))

    if gs.get("blocked_by_gate"):
        st.warning(f"**{gs['blocked_by_gate']} site(s) are asking for space while below "
                   f"{ex.OCCUPANCY_GATE:.0f}% occupancy.** There are "
                   f"{gs['free_where_blocked']:,} seats already free at those sites. The "
                   "request is not refused — it is answered with the space they hold.")

    if view_mode == "By site":
        st.markdown("##### What has to happen, and who raises it")
        if esc.empty:
            st.success("Nothing to escalate for this period.")
        else:
            for owner in esc["raise_with"].unique():
                sub = esc[esc["raise_with"] == owner]
                st.markdown(f"**{owner}** — {len(sub)} item(s), "
                            f"{int(sub['seats'].sum()):,} seats")
                st.dataframe(sub[["site", "seats", "action", "why"]],
                             width="stretch", hide_index=True)

        st.markdown("##### Site by site")
        cols_p = ["site", "occupancy_%", "required", "allocated", "incremental_need",
                  "available", "gate_met", "expansion_space", "after_expansion"]
        if "seats_returning" in pos.columns:
            cols_p += ["seats_returning", "net_need"]
        st.dataframe(pos[[c for c in cols_p if c in pos.columns]],
                     width="stretch", hide_index=True)

        st.markdown("##### Renovation programme")
        sched = ex.expansion_schedule(floors)
        if sched.empty:
            st.info("No renovation recorded in the floor inventory.")
        else:
            st.dataframe(sched, width="stretch", hide_index=True)
            late = sched[sched.get("lands", "") == "overdue"] if "lands" in sched else pd.DataFrame()
            if not late.empty:
                st.error(f"{len(late)} renovation(s) are past their expected date.")
    else:
        st.markdown("##### The renovation programme by year")
        yr = ex.expansion_by_year(floors)
        if yr.empty:
            st.info("No dated renovation to plot.")
        else:
            st.dataframe(yr, width="stretch", hide_index=True)
            st.bar_chart(yr.set_index(yr["year"].astype(str) + " " + yr["quarter"])["seats"],
                         height=260)
            st.caption("Seats unlocked by quarter, and the running total. This is the view a "
                       "budget holder wants — not which floor, but how much and when.")

    if not sunset_raw.empty:
        st.markdown("##### Programmes winding down")
        st.caption("The mirror of a deal not yet won: demand that is leaving. Just as easy to "
                   "forget until the floor empties.")
        st.dataframe(ex.sunset_effect(sunset_raw), width="stretch", hide_index=True)
        bysite = ex.sunset_by_site(sunset_raw)
        st.info(f"**{int(bysite['seats_returning'].sum()):,} seats** are due back across "
                f"{len(bysite)} site(s). Netting them off avoids asking for space that is "
                "about to free up.")

# ═══════════════════════════════════════════════ 4 · FLOOR MAP
with tab_map:
    st.caption("**Which client sits on which floor, and which seat?** Upload every plan for a "
               "site and demand is placed across its floors at once — which is how segregation "
               "actually works, since a client wanting its own space usually means its own floor.")

    if plan_file:
        payloads = tuple((f.name, f.getvalue()) for f in plan_file)
    else:
        import glob as _glob
        bundled = sorted(_glob.glob("data/plans/*.pdf"))
        payloads = tuple((_os.path.basename(p), open(p, "rb").read()) for p in bundled) \
            if bundled else ()
        if payloads:
            st.caption("No plans uploaded — showing the bundled surveyed site.")

    if not payloads:
        st.info("Upload one or more floor plans to see the map. Everything else in the app "
                "works without them.")
    else:
        library, unknown = build_library(payloads)
        if unknown:
            st.warning("Could not place " + str(len(unknown)) + " file(s):")
            st.dataframe(pd.DataFrame(unknown), width="stretch", hide_index=True)
        if not library:
            st.error("None of those files could be read as a floor plan.")
        else:
            rec = pl.reconcile(library, floors)
            sites_s = pl.surveyed_sites(library)
            c1, c2 = st.columns([1, 1])
            msite2 = c1.selectbox("Site", sites_s, key="map_site2")
            level2 = c2.radio("Colour by", ["Account", "LOB"], horizontal=True, key="map_lvl2")

            with st.expander(f"Plans loaded — {len(library)} floor(s) across "
                             f"{len(sites_s)} site(s)"):
                st.dataframe(rec, width="stretch", hide_index=True)
                off = rec[rec["agrees"] == "no"] if "agrees" in rec.columns else pd.DataFrame()
                if not off.empty:
                    st.warning("Some drawings disagree with the floor inventory. That usually "
                               "means the drawing is out of date, and it is worth resolving "
                               "before the plan is used.")

            site_plans = {k: v for k, v in library.items() if k[0] == msite2}
            seats_all = pl.combine(site_plans)
            allocz2 = sorted(seats_all.loc[seats_all["zone_type"].astype(str).str.lower()
                                           == "production", "zone"].unique())
            prod2 = int(seats_all["zone"].isin(allocz2).sum())

            labs2, sites2 = dr.site_options(demand)
            match = [l for l, sn in zip(labs2, sites2) if sn == msite2]
            demand_site = msite2 if match else None
            d1, d2 = st.columns([1, 1])
            if not match:
                pick = d1.selectbox("This site's demand is filed under", labs2, key="map_dem")
                demand_site = sites2[labs2.index(pick)]
            else:
                d1.caption(f"Demand for **{msite2}** found in the workbook.")
            mperiod2 = d2.select_slider("Period", options=dr.week_options(demand),
                                        value=dr.peak_week(demand, demand_site), key="map_per2")

            sl2 = dr.slice_week(demand, mperiod2, demand_site)
            if sl2.empty:
                st.info("No demand for that site in this period.")
            else:
                assigned2, blocks2, unplaced2 = pl.allocate_site(sl2, seats_all, allocz2)
                placed2 = int(assigned2["account"].notna().sum())
                need2 = int(sl2["seats"].sum())
                k2 = st.columns(4)
                k2[0].metric("Seats wanted", f"{need2:,}")
                k2[1].metric("Production seats", f"{prod2:,}")
                k2[2].metric("Placed", f"{placed2:,}")
                k2[3].metric("Left empty", f"{prod2 - placed2:,}")

                if not unplaced2.empty and int(unplaced2["short_by"].sum()) > 0:
                    st.error(f"{int(unplaced2['short_by'].sum()):,} seat(s) do not fit on the "
                             "surveyed floors at this site.")

                fs = pl.floor_summary(assigned2)
                sp = pl.account_spread(assigned2)
                t1, t2 = st.columns([1.35, 1])
                with t1:
                    st.markdown("###### Floor by floor")
                    st.dataframe(fs[["floor", "seats", "allocated", "empty", "full_%",
                                     "accounts", "shared"]],
                                 width="stretch", hide_index=True)
                with t2:
                    st.markdown("###### How each client landed")
                    st.dataframe(sp[["account", "seats", "floors", "verdict"]],
                                 width="stretch", hide_index=True)
                whole = int((sp["floors"] == 1).sum()) if not sp.empty else 0
                if whole:
                    st.success(f"{whole} client(s) have a floor to themselves at this site.")

                seg2 = sm.zone_security_report(assigned2)
                if seg2.empty:
                    st.success("Segregation clean — no zone is shared between two clients.")
                else:
                    st.warning(f"{len(seg2)} zone(s) shared by more than one client.")
                    st.dataframe(seg2, width="stretch", hide_index=True)

                if seat_model == "assigned":
                    st.markdown("###### Who sits at which desk")
                    if roster_raw is None or roster_raw.empty:
                        st.info("Add an employee roster — employee_id, account, site — to "
                                "place named people at named desks.")
                    else:
                        rprob = se.validate_roster(roster_raw)
                        if rprob:
                            st.error("Roster cannot be used:\n\n"
                                     + "\n".join(f"- {p}" for p in rprob))
                        else:
                            mine_r = roster_raw[roster_raw["site"] == msite2] \
                                if "site" in roster_raw.columns else roster_raw
                            asg = se.assign(mine_r, seats_all)
                            summ_a = se.assignment_summary(asg, seats_all)
                            ak = st.columns(4)
                            ak[0].metric("People", f"{summ_a.get('people', 0):,}")
                            ak[1].metric("Seated", f"{summ_a.get('seated', 0):,}")
                            ak[2].metric("Kept their desk", f"{summ_a.get('kept_existing', 0):,}")
                            ak[3].metric("Empty desks", f"{summ_a.get('empty_desks', 0):,}")
                            st.caption("People already sitting somewhere stay there. Moving "
                                       "somebody who does not need to move is the most "
                                       "expensive thing a seating plan can do.")
                            st.dataframe(asg, width="stretch", hide_index=True, height=260)
                            if summ_a.get("without_a_seat"):
                                st.error(f"{summ_a['without_a_seat']} person(s) have no desk "
                                         "on this site.")
                elif seat_model == "neighbourhood":
                    st.markdown("###### Team neighbourhoods")
                    nb = se.neighbourhoods(seats_all, demand, mperiod2)
                    if nb.empty:
                        st.info("No zones or no demand to match at this site.")
                    else:
                        st.dataframe(nb, width="stretch", hide_index=True)
                        waste = int(nb["spare_in_zone"].clip(lower=0).sum())
                        st.caption(f"A neighbourhood is taken whole, so {waste:,} seat(s) sit "
                                   "spare inside zones that are bigger than the team. That "
                                   "waste is the price of the model, and it should be visible "
                                   "rather than buried.")

                st.markdown("###### The floors")
                fl_list = pl.floors_for(library, msite2)
                ftabs = st.tabs([f"Floor {f}" for f in fl_list])
                for tab, fl in zip(ftabs, fl_list):
                    with tab:
                        key = next(k for k in site_plans if k[2] == fl)
                        this = assigned2[assigned2["floor"] == fl]
                        st.plotly_chart(
                            fr.plotly_map(this, background=site_plans[key].get("background"),
                                          extent=site_plans[key].get("extent"),
                                          level="account" if level2 == "Account" else "lob"),
                            width="stretch", key=f"map_{msite2}_{fl}")
                        png, pdf_b = fr.render_floor_map(
                            this, background=site_plans[key].get("background"),
                            extent=site_plans[key].get("extent"),
                            level="account" if level2 == "Account" else "lob",
                            title=f"{msite2} / {fl} — allocation by "
                                  f"{'client' if level2 == 'Account' else 'line of business'}",
                            subtitle=f"{mperiod2} · {len(this)} seats on this floor")
                        e1, e2 = st.columns(2)
                        e1.download_button("🖼 PNG", png, key=f"png_{fl}",
                                           file_name=f"{msite2}_{fl}_{mperiod2}.png",
                                           mime="image/png", width="stretch")
                        e2.download_button("📄 PDF", pdf_b, key=f"pdf_{fl}",
                                           file_name=f"{msite2}_{fl}_{mperiod2}.pdf",
                                           mime="application/pdf", width="stretch")

                st.download_button("📊 Seat register, whole site (Excel)",
                                   _xlsx({"Seats": assigned2, "By floor": fs,
                                          "By client": sp, "Segregation": seg2}),
                                   file_name=f"Floorcast_{msite2}_{mperiod2}.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument."
                                        "spreadsheetml.sheet")

# ═══════════════════════════════════════════════ 5 · SCENARIOS
with tab_scen:
    st.caption("**What are our options when it does not fit?**")
    st.markdown("#### Release trapped seats")
    st.caption("What the estate looks like if trapped seats are recovered — a partition "
               "move, an IT rollout, a relocation, or finishing a renovation.")
    tb = es.trapped_breakdown(floors)
    default = tb.loc[tb["releasable"], "trapped_reason"].tolist()
    s1, s2 = st.columns([2, 1])
    reasons = s1.multiselect("Reasons to release", tb["trapped_reason"].tolist(),
                             default=default)
    frac = s2.slider("Proportion recovered", 0.0, 1.0, 0.5, 0.05)
    after = es.release_trapped(floors, reasons, frac)
    b, a = es.estate_totals(floors), es.estate_totals(after)
    k = st.columns(4)
    k[0].metric("Available before", f"{b['available']:,}")
    k[1].metric("Available after", f"{a['available']:,}",
                delta=f"{a['available'] - b['available']:+,}")
    k[2].metric("Trapped after", f"{a['trapped']:,}",
                delta=f"{a['trapped'] - b['trapped']:+,}", delta_color="inverse")
    k[3].metric("Equivalent to", f"{(a['available'] - b['available']) / max(b['total_seats'], 1) * 100:.1f}%",
                help="of the total estate, recovered without new space")
    st.dataframe(after.loc[after["released"] > 0,
                           ["site", "floor_id", "trapped_reason", "released",
                            "available", "trapped"]],
                 width="stretch", hide_index=True)

    st.markdown("#### Seats shared across shifts")
    st.caption("Two accounts working opposite hours can use the same desk — provided their "
               "systems can be handed over and no rule keeps them apart.")
    if shifts_raw is None or shifts_raw.empty:
        st.info("Add a shift pattern file — account, site, shift, start_hour, end_hour, "
                "system_locked — to see which accounts could share seats.")
    else:
        probs_sh = sg.validate(shifts_raw)
        if probs_sh:
            st.error("Shift file cannot be used:\n\n" + "\n".join(f"- {p}" for p in probs_sh))
        else:
            sp_req = {}
            pr = sg.pairs(shifts_raw, rules=RULES, space_requirements=sp_req)
            sv = sg.savings(pr, demand, dr.peak_week(demand))
            st.info(sg.note(sv))
            if not sv.empty:
                st.dataframe(sv, width="stretch", hide_index=True)
            with st.expander("Pairs that cannot share, and why"):
                blocked = pr[pr["can_share"] == "no"]
                st.dataframe(blocked[["site", "account_a", "account_b", "overlap_hours",
                                      "why_not"]] if not blocked.empty else
                             pd.DataFrame({"note": ["Every pair at every site could share."]}),
                             width="stretch", hide_index=True)

    st.markdown("#### Deals not yet won")
    st.caption("Business sales is chasing but has not signed. The likely view is what you plan "
               "against; all of them is the stress test — the estate should have an answer if "
               "every deal comes in.")
    p1, p2 = st.columns([2, 1])
    mode = p1.radio("Count these deals as", ["exclude", "weighted", "full"], index=1,
                    horizontal=True,
                    format_func=lambda m: {"exclude": "Signed business only",
                                           "weighted": "Weighted by how likely",
                                           "full": "Every deal we are chasing"}[m])
    ratio = p2.number_input("Seat ratio for these deals", 1.0, 3.0, 1.2, 0.1)
    ps = dr.pipeline_seats(pipe, mode, ratio)
    pv = ps.groupby("site")["seats"].sum().rename("seats_if_won").reset_index()
    cap = es.rollup(floors, "site")[["site", "available", "trapped", "expansion_space"]]
    comp = cap.merge(pv, left_on="site", right_on="site", how="outer").fillna(0)
    comp["gap_vs_available"] = comp["available"] - comp["seats_if_won"]
    st.dataframe(comp, width="stretch", hide_index=True)
    tight = comp[comp["gap_vs_available"] < 0]
    if not tight.empty:
        st.warning("These deals would need more seats than are free at: "
                   + ", ".join(f"{r['site']} (short {int(-r['gap_vs_available'])})"
                               for _, r in tight.iterrows())
                   + ". Trapped seats and expansion space are the first places to look.")
    else:
        st.success("Every site could absorb these deals on the seats free today.")

# ═══════════════════════════════════════════════ 7 · DATA QUALITY
with tab_qa:
    st.caption("**Is the plan built on numbers that add up, and how much of it has arrived?** "
               "Point this at the full planning workbook — one tab per geo — rather than a "
               "single-sheet extract.")
    q_file = st.file_uploader("Planning workbook (multi-tab Excel)", type=["xlsx", "xls"],
                              key="qa_file",
                              help="The collection file with a tab per country and a block of "
                                   "metric rows per programme")

    @st.cache_data(show_spinner="Reading the workbook…")
    def load_os(data):
        return os1.read_workbook(io.BytesIO(data))

    try:
        src = load_os(q_file.getvalue()) if q_file else \
            load_os(open("data/sample_one_source.xlsx", "rb").read())
    except Exception as exc:
        st.error(f"Could not read that workbook: {exc}")
        src = None

    if src is None or src["programmes"].empty:
        st.info("No programme blocks found. This view expects the collection format — "
                "Site / Account / Information / Data columns, with a block of metric rows "
                "starting at 'Total TMs' for each programme.")
    else:
        progs, longd = src["programmes"], src["long"]
        if not q_file:
            st.caption("Running on a bundled sample in the collection format.")
        issues = os1.quality_report(longd)
        comp = os1.completion(progs)

        k = st.columns(4)
        k[0].metric("Programmes", f"{len(progs):,}")
        k[1].metric("Geos", progs["geo"].nunique())
        done = int(progs["status"].eq("Updated").sum())
        k[2].metric("Collected", f"{done / max(len(progs), 1) * 100:.0f}%",
                    delta=f"{len(progs) - done} outstanding", delta_color="inverse")
        n_high = int((issues["severity"] == "High").sum()) if not issues.empty else 0
        k[3].metric("Programmes to check", f"{issues['programme'].nunique() if not issues.empty else 0:,}",
                    delta=f"{n_high} high severity" if n_high else None, delta_color="inverse")

        st.warning(os1.coverage_note(progs, longd))

        qa1, qa2, qa3 = st.tabs(["What does not add up", "Collection status", "Who to chase"])
        with qa1:
            if issues.empty:
                st.success("Every programme passes the checks.")
            else:
                st.dataframe(os1.quality_summary(issues, progs), width="stretch", hide_index=True)
                st.caption("Seats Required is taken from the workbook as authoritative. Where a "
                           "recomputation disagrees it is flagged, never overwritten — the "
                           "guidance is that the calculation belongs in the cap plans.")
                sev = st.multiselect("Severity", ["High", "Medium", "Low"],
                                     default=["High", "Medium"], key="qa_sev")
                view = issues[issues["severity"].isin(sev)]
                st.dataframe(view[["geo", "site", "account", "programme", "issue",
                                   "severity", "worst_period", "detail"]],
                             width="stretch", hide_index=True, height=320)
                st.download_button("📊 Issue list (Excel)",
                                   _xlsx({"Issues": issues,
                                          "Summary": os1.quality_summary(issues, progs)}),
                                   file_name="Floorcast_DataQuality.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument."
                                        "spreadsheetml.sheet")
        with qa2:
            st.dataframe(comp, width="stretch", hide_index=True)
            if not comp.empty:
                st.bar_chart(comp.set_index("geo")["completion_%"], height=240)
                worst = comp.iloc[0]
                st.caption(f"Lowest is {worst['geo']} at {worst['completion_%']}% "
                           f"({int(worst['not_updated'])} of {int(worst['programmes'])} "
                           "programmes outstanding).")
        with qa3:
            chase = os1.chase_list(progs)
            if chase.empty:
                st.success("Nothing outstanding.")
            else:
                st.dataframe(chase, width="stretch", hide_index=True)
                st.download_button("📋 Chase list (CSV)", chase.to_csv(index=False).encode(),
                                   file_name="not_updated.csv", mime="text/csv")
                st.caption("Generated from the workbook itself, so it cannot drift out of step "
                           "with the data the way a maintained list does.")
