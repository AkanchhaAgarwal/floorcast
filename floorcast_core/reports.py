"""
Floorcast — Report Generator
Builds multi-sheet Excel workbooks (in-memory) for download from the app.

Report 1 (roster mode): Summary · Capacity Plan · Seat Demand · Bay Plan
Report 2 (forecast mode): Summary · Capacity Plan · Shift Schedule · Floor Plan
"""

import io
import math
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEAL = "0F6B6B"
GOLD = "C9962E"
LIGHT = "E6F2F2"
FONT = "Arial"
PROG_FILLS = ["0F6B6B", "C9962E", "7C3AED", "DC2626", "2563EB", "059669"]

_thin = Side(style="thin", color="D1D5DB")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", start_color=TEAL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def _write_df(ws, df, start_row=1, title=None):
    r = start_row
    if title:
        ws.cell(row=r, column=1, value=title).font = Font(name=FONT, bold=True, size=12, color=TEAL)
        r += 2
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=r, column=j, value=str(col))
    _style_header(ws, r, len(df.columns))
    for i, row in enumerate(df.itertuples(index=False), start=r + 1):
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = Font(name=FONT, size=10)
            c.border = BORDER
    for j, col in enumerate(df.columns, start=1):
        width = max(12, min(28, int(df[col].astype(str).str.len().max() or 10) + 4, len(str(col)) + 6))
        ws.column_dimensions[get_column_letter(j)].width = width
    return r + len(df) + 1


def _title_sheet(ws, title, lines):
    ws.cell(row=1, column=1, value="Floorcast").font = Font(name=FONT, bold=True, size=20, color=TEAL)
    ws.cell(row=2, column=1, value=title).font = Font(name=FONT, italic=True, size=12, color=GOLD)
    r = 4
    for k, v in lines:
        ws.cell(row=r, column=1, value=k).font = Font(name=FONT, bold=True, size=10)
        ws.cell(row=r, column=2, value=v).font = Font(name=FONT, size=10)
        r += 1
    ws.cell(row=r + 1, column=1,
            value="Planning draft — layouts must be verified by a licensed architect "
                  "(fire egress, travel distances, occupancy limits) before implementation."
            ).font = Font(name=FONT, italic=True, size=9, color="6B7280")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 46


# ────────────────────────────────────────────── Report 1: roster mode
def roster_report(roster, demand, summary, alloc, diag, day) -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    _title_sheet(ws, "Seat Planner report (roster mode)", [
        ("Generated for date", str(day)),
        ("Rostered agents", int(roster["agent_id"].nunique())),
        ("Office agents", int(roster[roster.work_mode.eq("Office")]["agent_id"].nunique())),
        ("Peak seats required", int(summary["peak_seats"].sum())),
        ("Solver status", diag.get("status", "")),
        ("Bays used", diag.get("bays_used", "")),
    ])

    ws2 = wb.create_sheet("Capacity Plan")
    _write_df(ws2, summary.rename(columns={
        "program": "Program", "office_headcount": "Office headcount",
        "peak_seats": "Peak seats", "seat_sharing_ratio": "Seat-sharing ratio"}),
        title="Per-program capacity plan")

    ws3 = wb.create_sheet("Seat Demand")
    piv = demand[demand["date"] == day].pivot(index="time", columns="program",
                                              values="seats_needed").reset_index()
    piv.columns.name = None
    end = _write_df(ws3, piv.rename(columns={"time": "Interval"}),
                    title=f"Interval seat demand — {day}")
    # conditional-style peak rows
    for col in range(2, piv.shape[1] + 1):
        vals = [ws3.cell(row=r, column=col).value for r in range(4, 4 + len(piv))]
        peak = max(v for v in vals if v is not None)
        for r in range(4, 4 + len(piv)):
            if ws3.cell(row=r, column=col).value == peak:
                ws3.cell(row=r, column=col).fill = PatternFill("solid", start_color="FCE9C8")

    ws4 = wb.create_sheet("Bay Plan")
    if alloc is not None and not alloc.empty:
        a = alloc.rename(columns={"program": "Program", "bay": "Bay", "floor": "Floor",
                                  "seats": "Seats allocated", "bay_capacity": "Bay capacity"}).copy()
        a["Utilization %"] = (a["Seats allocated"] / a["Bay capacity"] * 100).round(1)
        _write_df(ws4, a, title="Optimized bay allocation")
    else:
        ws4.cell(row=1, column=1, value="Allocation infeasible — demand exceeds capacity "
                 "under current constraints.").font = Font(name=FONT, bold=True, color="DC2626")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ────────────────────────────────────────────── Report 2: forecast mode
def forecast_report(res: dict, pods_needed: dict, grid, part_ft, month,
                    sl_pct, shrink, pod_seats, times) -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    _title_sheet(ws, f"Forecast → Floor report — month {month}", [
        ("Service level target", f"{int(sl_pct*100)}% "),
        ("Shrinkage", f"{int(shrink*100)}%"),
        ("Programs", len(res)),
        ("Total scheduled agents", sum(d["scheduled"] for d in res.values())),
        ("Total headcount (after shrinkage)", sum(d["total_hc"] for d in res.values())),
        ("Total peak seats", sum(d["peak_seats"] for d in res.values())),
        ("Pods required", sum(pods_needed.values())),
        ("Partition length (linear ft)", None if part_ft is None else round(part_ft)),
    ])

    ws2 = wb.create_sheet("Capacity Plan")
    cap = pd.DataFrame([{
        "Program": p, "Channel": d["channel"],
        "Peak agents on-seat (Erlang C)": int(d["required"].max()),
        "Scheduled agents (shift mix)": d["scheduled"],
        "Headcount after shrinkage": d["total_hc"],
        "Peak seats (after WFH)": d["peak_seats"],
        "Pods required": pods_needed[p],
    } for p, d in res.items()])
    _write_df(ws2, cap, title=f"Capacity plan — month {month}")

    ws3 = wb.create_sheet("Shift Schedule")
    rows = []
    for p, d in res.items():
        for start, n in sorted(d["mix"].items()):
            h = int(start); m = int((start - h) * 60)
            eh = (start + 9) % 24; eh_i = int(eh); em = int((eh - eh_i) * 60)
            rows.append({"Program": p, "Shift": f"{h:02d}:{m:02d}-{eh_i:02d}:{em:02d}",
                         "Agents": n})
    _write_df(ws3, pd.DataFrame(rows), title="Optimized shift mix (set-covering solution)")

    ws4 = wb.create_sheet("Interval Coverage")
    cov = pd.DataFrame({"Interval": times})
    for p, d in res.items():
        cov[f"{p} required"] = d["required"]
        cov[f"{p} scheduled"] = d["coverage"]
        cov[f"{p} on-site"] = d["onsite"]
    _write_df(ws4, cov, title="30-minute interval coverage")

    ws5 = wb.create_sheet("Floor Plan")
    ws5.cell(row=1, column=1, value="Floor plan — pod grid (each cell = "
             f"{pod_seats} seats)").font = Font(name=FONT, bold=True, size=12, color=TEAL)
    if grid is None:
        ws5.cell(row=3, column=1, value="INFEASIBLE — pods required exceed floor capacity."
                 ).font = Font(name=FONT, bold=True, color="DC2626")
    else:
        progs = list(res.keys())
        fills = {p: PROG_FILLS[i % len(PROG_FILLS)] for i, p in enumerate(progs)}
        nrows, ncols = grid.shape
        for r in range(nrows):
            ws5.row_dimensions[r + 3].height = 26
            for c in range(ncols):
                cell = ws5.cell(row=r + 3, column=c + 2)
                p = grid[r][c]
                cell.border = BORDER
                if p:
                    cell.value = p[:4]
                    cell.fill = PatternFill("solid", start_color=fills[p])
                    cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=8)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.fill = PatternFill("solid", start_color="F3F4F6")
        for c in range(ncols):
            ws5.column_dimensions[get_column_letter(c + 2)].width = 7
        lr = nrows + 4
        ws5.cell(row=lr, column=2, value="Legend:").font = Font(name=FONT, bold=True, size=9)
        for i, p in enumerate(progs):
            cell = ws5.cell(row=lr + 1 + i, column=2, value=p)
            cell.fill = PatternFill("solid", start_color=fills[p])
            cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=9)
        ws5.cell(row=lr + len(progs) + 2, column=2,
                 value="Draft only — architect certification required for exits/egress."
                 ).font = Font(name=FONT, italic=True, size=9, color="6B7280")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
